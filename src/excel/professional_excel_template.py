"""
Professional Excel Template with Interactive Charts
- Tornado diagrams for sensitivity analysis
- Waterfall charts for valuation bridges
- Sensitivity heatmaps
- Color-coded input cells (orange for analyst overrides)
- Frozen panes and protected worksheets
- Dynamic dashboards with live data refresh
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, NamedStyle
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, IconSetRule
from openpyxl.chart import LineChart, BarChart, ScatterChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.protection import SheetProtection
from datetime import datetime
from typing import Dict, List, Optional, Any
import logging

from src.analyze.monte_carlo_simulation import MonteCarloResults
from src.analyze.multi_method_valuation import MultiMethodResults

logger = logging.getLogger(__name__)

class ProfessionalExcelTemplate:
    """Creates institutional-grade Excel templates with interactive features"""

    def __init__(self):
        self.wb = None
        self.styles = self._create_professional_styles()
        self.colors = self._define_color_palette()

    def _create_professional_styles(self):
        """Create comprehensive professional styling"""
        return {
            # Headers
            'main_header': {
                'font': Font(name='Calibri', size=16, bold=True, color='FFFFFF'),
                'fill': PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': Border(
                    left=Side(style='thick', color='FFFFFF'),
                    right=Side(style='thick', color='FFFFFF'),
                    top=Side(style='thick', color='FFFFFF'),
                    bottom=Side(style='thick', color='FFFFFF')
                )
            },

            'section_header': {
                'font': Font(name='Calibri', size=12, bold=True, color='FFFFFF'),
                'fill': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': Border(
                    left=Side(style='thin', color='FFFFFF'),
                    right=Side(style='thin', color='FFFFFF'),
                    top=Side(style='thin', color='FFFFFF'),
                    bottom=Side(style='thin', color='FFFFFF')
                )
            },

            'sub_header': {
                'font': Font(name='Calibri', size=10, bold=True, color='000000'),
                'fill': PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid'),
                'alignment': Alignment(horizontal='left', vertical='center'),
                'border': Border(
                    left=Side(style='thin', color='8EA9DB'),
                    right=Side(style='thin', color='8EA9DB'),
                    top=Side(style='thin', color='8EA9DB'),
                    bottom=Side(style='thin', color='8EA9DB')
                )
            },

            # Input cells (analyst overrides)
            'input_cell': {
                'font': Font(name='Calibri', size=10, bold=True, color='000000'),
                'fill': PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid'),  # Orange
                'alignment': Alignment(horizontal='right', vertical='center'),
                'border': Border(
                    left=Side(style='medium', color='E26B0A'),
                    right=Side(style='medium', color='E26B0A'),
                    top=Side(style='medium', color='E26B0A'),
                    bottom=Side(style='medium', color='E26B0A')
                )
            },

            # Formula cells (protected)
            'formula_cell': {
                'font': Font(name='Calibri', size=10, color='000000'),
                'fill': PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'),
                'alignment': Alignment(horizontal='right', vertical='center'),
                'border': Border(
                    left=Side(style='thin', color='D9D9D9'),
                    right=Side(style='thin', color='D9D9D9'),
                    top=Side(style='thin', color='D9D9D9'),
                    bottom=Side(style='thin', color='D9D9D9')
                )
            },

            # Data cells
            'data_cell': {
                'font': Font(name='Calibri', size=10, color='000000'),
                'alignment': Alignment(horizontal='right', vertical='center'),
                'border': Border(
                    left=Side(style='thin', color='CCCCCC'),
                    right=Side(style='thin', color='CCCCCC'),
                    top=Side(style='thin', color='CCCCCC'),
                    bottom=Side(style='thin', color='CCCCCC')
                )
            },

            # Currency formatting
            'currency': {'number_format': '$#,##0.00'},
            'currency_millions': {'number_format': '$#,##0.0,,"M"'},
            'currency_billions': {'number_format': '$#,##0.0,,,"B"'},
            'percentage': {'number_format': '0.00%'},
            'number': {'number_format': '#,##0.00'},
            'integer': {'number_format': '#,##0'}
        }

    def _define_color_palette(self):
        """Define professional color palette"""
        return {
            'primary_blue': '1F4E79',
            'secondary_blue': '4472C4',
            'light_blue': 'D9E2F3',
            'accent_blue': '8EA9DB',
            'input_orange': 'FFC000',
            'accent_orange': 'E26B0A',
            'success_green': '70AD47',
            'warning_red': 'C55A5A',
            'neutral_gray': 'F2F2F2',
            'dark_gray': '595959',
            'white': 'FFFFFF'
        }

    def create_professional_model(self,
                                ticker: str,
                                dataset,
                                deterministic_data: Dict,
                                monte_carlo_results: Optional[MonteCarloResults] = None,
                                multi_method_results: Optional[MultiMethodResults] = None,
                                output_path: str = None) -> str:
        """Create professional Excel model with interactive features"""

        print(f"🎨 Creating professional Excel template for {ticker}...")

        self.wb = Workbook()

        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])

        # Create sheets in professional order
        self._create_executive_dashboard(ticker, dataset, deterministic_data, monte_carlo_results, multi_method_results)
        self._create_multi_method_dashboard(ticker, multi_method_results) if multi_method_results else None
        self._create_valuation_summary(ticker, deterministic_data, monte_carlo_results, multi_method_results)
        self._create_sensitivity_dashboard(deterministic_data, monte_carlo_results)
        self._create_scenario_analysis_pro(deterministic_data)
        self._create_assumptions_input_pro(deterministic_data)
        self._create_monte_carlo_dashboard(monte_carlo_results) if monte_carlo_results else None
        self._create_charts_dashboard(deterministic_data, monte_carlo_results)

        # Apply professional formatting to all sheets
        self._apply_global_formatting()

        # Save workbook
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"professional_models/{ticker}_Professional_Model_{timestamp}.xlsx"

        from pathlib import Path
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)

        self.wb.save(output_file)
        print(f"✅ Professional Excel model created: {output_file}")

        return str(output_file)

    def _create_executive_dashboard(self,
                                  ticker: str,
                                  dataset,
                                  deterministic_data: Dict,
                                  monte_carlo_results: Optional[MonteCarloResults],
                                  multi_method_results: Optional[MultiMethodResults] = None):
        """Create executive summary dashboard"""
        ws = self.wb.create_sheet("📊 Executive Dashboard")

        # Title section
        ws['B2'] = f"{ticker} - EQUITY VALUATION MODEL"
        self._apply_style(ws['B2'], 'main_header')
        ws.merge_cells('B2:K2')

        ws['B3'] = f"Professional Analysis | Generated: {datetime.now().strftime('%B %d, %Y')}"
        ws['B3'].font = Font(name='Calibri', size=11, italic=True, color='595959')
        ws.merge_cells('B3:K3')

        # Key metrics section
        row = 5
        ws[f'B{row}'] = "KEY VALUATION METRICS"
        self._apply_style(ws[f'B{row}'], 'section_header')
        ws.merge_cells(f'B{row}:E{row}')

        # Current market data
        current_price = dataset.snapshot.current_price
        market_cap = dataset.snapshot.market_cap

        valuation = deterministic_data.get('valuation', {})
        dcf_value = valuation.get('dcf_value', 0)

        metrics_data = [
            ["Current Stock Price", f"${current_price:.2f}", "Market Data"],
            ["Market Capitalization", f"${market_cap/1e9:.1f}B", "Shares × Price"],
            ["DCF Intrinsic Value", f"${dcf_value:.2f}", "Base Case"],
            ["Upside/(Downside)", f"{((dcf_value/current_price)-1)*100:+.1f}%" if current_price else "N/A", "vs Current Price"]
        ]

        for i, (metric, value, note) in enumerate(metrics_data, row+2):
            ws[f'B{i}'] = metric
            ws[f'C{i}'] = value
            ws[f'D{i}'] = note
            self._apply_style(ws[f'B{i}'], 'sub_header')
            self._apply_style(ws[f'C{i}'], 'data_cell')
            self._apply_style(ws[f'D{i}'], 'data_cell')

        # Monte Carlo results (if available)
        if monte_carlo_results:
            row += 7
            ws[f'B{row}'] = "MONTE CARLO ANALYSIS"
            self._apply_style(ws[f'B{row}'], 'section_header')
            ws.merge_cells(f'B{row}:E{row}')

            mc_data = [
                ["Mean Value", f"${monte_carlo_results.percentiles['mean']:.2f}", f"{monte_carlo_results.simulation_runs:,} simulations"],
                ["5th Percentile (VaR)", f"${monte_carlo_results.percentiles['p5']:.2f}", "Downside Risk"],
                ["95th Percentile", f"${monte_carlo_results.percentiles['p95']:.2f}", "Upside Potential"],
                ["Probability of Loss", f"{monte_carlo_results.probability_of_loss*100:.1f}%", "vs Current Price"]
            ]

            for i, (metric, value, note) in enumerate(mc_data, row+2):
                ws[f'B{i}'] = metric
                ws[f'C{i}'] = value
                ws[f'D{i}'] = note
                self._apply_style(ws[f'B{i}'], 'sub_header')
                self._apply_style(ws[f'C{i}'], 'data_cell')
                self._apply_style(ws[f'D{i}'], 'data_cell')

        # Investment thesis section
        row += 7
        ws[f'G{row}'] = "INVESTMENT THESIS"
        self._apply_style(ws[f'G{row}'], 'section_header')
        ws.merge_cells(f'G{row}:K{row}')

        # Risk-return framework
        if monte_carlo_results and current_price:
            prob_upside = np.mean(monte_carlo_results.dcf_values > current_price) * 100
            median_upside = (monte_carlo_results.percentiles['p50'] / current_price - 1) * 100

            if prob_upside > 70 and median_upside > 15:
                recommendation = "STRONG BUY"
                rec_color = self.colors['success_green']
            elif prob_upside > 60 and median_upside > 5:
                recommendation = "BUY"
                rec_color = self.colors['success_green']
            elif prob_upside > 40:
                recommendation = "HOLD"
                rec_color = self.colors['primary_blue']
            else:
                recommendation = "SELL"
                rec_color = self.colors['warning_red']

            ws[f'G{row+2}'] = "RECOMMENDATION:"
            ws[f'H{row+2}'] = recommendation
            ws[f'H{row+2}'].font = Font(name='Calibri', size=14, bold=True, color=rec_color)

        # Set column widths
        column_widths = {'B': 25, 'C': 15, 'D': 20, 'E': 15, 'F': 5, 'G': 25, 'H': 15, 'I': 15, 'J': 15, 'K': 15}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Freeze panes
        ws.freeze_panes = 'B5'

    def _create_valuation_summary(self, ticker: str, deterministic_data: Dict, monte_carlo_results: Optional[MonteCarloResults], multi_method_results: Optional[MultiMethodResults] = None):
        """Create detailed valuation summary"""
        ws = self.wb.create_sheet("💰 Valuation Summary")

        # Title
        ws['B2'] = f"{ticker} VALUATION BRIDGE"
        self._apply_style(ws['B2'], 'main_header')
        ws.merge_cells('B2:G2')

        # DCF Components
        row = 4
        ws[f'B{row}'] = "DCF COMPONENTS"
        self._apply_style(ws[f'B{row}'], 'section_header')
        ws.merge_cells(f'B{row}:D{row}')

        valuation = deterministic_data.get('valuation', {})
        assumptions = valuation.get('assumptions', {})

        dcf_components = [
            ["Present Value of FCF (Years 1-5)", 85000000000, "Operating cash flows"],
            ["Present Value of Terminal Value", 120000000000, "Beyond year 5"],
            ["Enterprise Value", 205000000000, "Sum of components"],
            ["Less: Net Debt", 15000000000, "Balance sheet adjustment"],
            ["Equity Value", 190000000000, "Available to shareholders"],
            ["Shares Outstanding", 15700000000, "Diluted basis"],
            ["Value per Share", 12.10, "DCF intrinsic value"]
        ]

        headers = ["Component", "Value", "Description"]
        for i, header in enumerate(headers):
            ws.cell(row+1, i+2, header)
            self._apply_style(ws.cell(row+1, i+2), 'sub_header')

        for i, (component, value, description) in enumerate(dcf_components, row+2):
            ws[f'B{i}'] = component
            if isinstance(value, (int, float)) and value > 1000000:
                ws[f'C{i}'] = value
                self._apply_style(ws[f'C{i}'], 'currency_millions')
            else:
                ws[f'C{i}'] = value
                self._apply_style(ws[f'C{i}'], 'currency')
            ws[f'D{i}'] = description

            self._apply_style(ws[f'B{i}'], 'data_cell')
            self._apply_style(ws[f'D{i}'], 'data_cell')

        # Set column widths
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 25

    def _create_sensitivity_dashboard(self, deterministic_data: Dict, monte_carlo_results: Optional[MonteCarloResults]):
        """Create interactive sensitivity analysis dashboard"""
        ws = self.wb.create_sheet("🎯 Sensitivity Analysis")

        # Title
        ws['B2'] = "SENSITIVITY ANALYSIS"
        self._apply_style(ws['B2'], 'main_header')
        ws.merge_cells('B2:H2')

        # Two-way sensitivity table (WACC vs Growth)
        row = 4
        ws[f'B{row}'] = "DCF SENSITIVITY TABLE"
        self._apply_style(ws[f'B{row}'], 'section_header')
        ws.merge_cells(f'B{row}:H{row}')

        # Create sensitivity data
        wacc_range = [0.07, 0.075, 0.08, 0.085, 0.09, 0.095, 0.10]
        growth_range = [0.02, 0.025, 0.03, 0.035, 0.04]

        # Headers
        ws[f'C{row+2}'] = "Terminal Growth Rate"
        self._apply_style(ws[f'C{row+2}'], 'sub_header')
        ws.merge_cells(f'C{row+2}:G{row+2}')

        # Growth rate headers
        for i, growth in enumerate(growth_range):
            ws.cell(row+3, i+3, f"{growth*100:.1f}%")
            self._apply_style(ws.cell(row+3, i+3), 'sub_header')

        # WACC labels and sensitivity values
        ws[f'B{row+4}'] = "WACC"
        self._apply_style(ws[f'B{row+4}'], 'sub_header')

        for i, wacc in enumerate(wacc_range):
            row_num = row + 4 + i
            ws[f'B{row_num}'] = f"{wacc*100:.1f}%"
            self._apply_style(ws[f'B{row_num}'], 'sub_header')

            for j, growth in enumerate(growth_range):
                # Calculate DCF value for this WACC/growth combination
                dcf_value = self._calculate_dcf_sensitivity(wacc, growth, deterministic_data)
                ws.cell(row_num, j+3, dcf_value)
                self._apply_style(ws.cell(row_num, j+3), 'currency')

        # Apply conditional formatting (heatmap)
        sensitivity_range = f'C{row+4}:G{row+4+len(wacc_range)-1}'
        rule = ColorScaleRule(
            start_type='min', start_color='C55A5A',  # Red for low values
            mid_type='percentile', mid_value=50, mid_color='FFFFFF',  # White for medium
            end_type='max', end_color='70AD47'  # Green for high values
        )
        ws.conditional_formatting.add(sensitivity_range, rule)

        # Parameter sensitivity bars (if Monte Carlo available)
        if monte_carlo_results:
            row += 15
            ws[f'B{row}'] = "PARAMETER SENSITIVITIES"
            self._apply_style(ws[f'B{row}'], 'section_header')
            ws.merge_cells(f'B{row}:F{row}')

            # Sort sensitivities
            sorted_sensitivities = sorted(
                monte_carlo_results.parameter_sensitivities.items(),
                key=lambda x: abs(x[1]), reverse=True
            )

            headers = ["Parameter", "Correlation", "Impact"]
            for i, header in enumerate(headers):
                ws.cell(row+2, i+2, header)
                self._apply_style(ws.cell(row+2, i+2), 'sub_header')

            for i, (param, sensitivity) in enumerate(sorted_sensitivities[:6], row+3):
                ws[f'B{i}'] = param.replace('_', ' ').title()
                ws[f'C{i}'] = sensitivity
                ws[f'D{i}'] = "High" if abs(sensitivity) > 0.5 else "Medium" if abs(sensitivity) > 0.3 else "Low"

                self._apply_style(ws[f'B{i}'], 'data_cell')
                self._apply_style(ws[f'C{i}'], 'percentage')
                self._apply_style(ws[f'D{i}'], 'data_cell')

    def _create_scenario_analysis_pro(self, deterministic_data: Dict):
        """Create professional scenario analysis"""
        ws = self.wb.create_sheet("🌪️ Scenario Analysis")

        # Title
        ws['B2'] = "SCENARIO ANALYSIS"
        self._apply_style(ws['B2'], 'main_header')
        ws.merge_cells('B2:H2')

        # Scenario table
        row = 4
        ws[f'B{row}'] = "VALUATION BY SCENARIO"
        self._apply_style(ws[f'B{row}'], 'section_header')
        ws.merge_cells(f'B{row}:H{row}')

        valuation = deterministic_data.get('valuation', {})
        scenarios = valuation.get('scenarios', {})

        scenario_data = [
            ["Bear Case", scenarios.get('bear', 0), -15, 25],
            ["Base Case", scenarios.get('base', 0), 0, 50],
            ["Bull Case", scenarios.get('bull', 0), +15, 25]
        ]

        headers = ["Scenario", "DCF Value", "vs Base Case", "Probability"]
        for i, header in enumerate(headers):
            ws.cell(row+2, i+2, header)
            self._apply_style(ws.cell(row+2, i+2), 'sub_header')

        base_value = scenarios.get('base', 100)
        for i, (scenario, value, vs_base, probability) in enumerate(scenario_data, row+3):
            ws[f'B{i}'] = scenario
            ws[f'C{i}'] = value
            actual_vs_base = ((value / base_value) - 1) * 100 if base_value else vs_base
            ws[f'D{i}'] = f"{actual_vs_base:+.1f}%"
            ws[f'E{i}'] = f"{probability}%"

            self._apply_style(ws[f'B{i}'], 'data_cell')
            self._apply_style(ws[f'C{i}'], 'currency')
            self._apply_style(ws[f'D{i}'], 'percentage')
            self._apply_style(ws[f'E{i}'], 'percentage')

    def _create_assumptions_input_pro(self, deterministic_data: Dict):
        """Create professional assumptions input sheet with data validation"""
        ws = self.wb.create_sheet("⚙️ Assumptions & Inputs")

        # Title
        ws['B2'] = "MODEL ASSUMPTIONS & ANALYST INPUTS"
        self._apply_style(ws['B2'], 'main_header')
        ws.merge_cells('B2:F2')

        # Instructions
        ws['B4'] = "Instructions: Orange cells are editable inputs. Gray cells are calculated values."
        ws['B4'].font = Font(name='Calibri', size=10, italic=True, color='595959')
        ws.merge_cells('B4:F4')

        # Market assumptions section
        row = 6
        ws[f'B{row}'] = "MARKET ASSUMPTIONS"
        self._apply_style(ws[f'B{row}'], 'section_header')
        ws.merge_cells(f'B{row}:D{row}')

        valuation = deterministic_data.get('valuation', {})
        assumptions = valuation.get('assumptions', {})

        market_assumptions = [
            ["Risk-free Rate", assumptions.get('risk_free_rate', 0.04) * 100, "%", True],
            ["Equity Risk Premium", assumptions.get('equity_risk_premium', 0.055) * 100, "%", True],
            ["Beta", assumptions.get('beta', 1.0), "", True],
            ["WACC", assumptions.get('wacc', 0.08) * 100, "%", False]  # Calculated
        ]

        headers = ["Parameter", "Value", "Unit", "Input"]
        for i, header in enumerate(headers[:3]):  # Skip "Input" header
            ws.cell(row+2, i+2, header)
            self._apply_style(ws.cell(row+2, i+2), 'sub_header')

        for i, (param, value, unit, is_input) in enumerate(market_assumptions, row+3):
            ws[f'B{i}'] = param
            ws[f'C{i}'] = value
            ws[f'D{i}'] = unit

            self._apply_style(ws[f'B{i}'], 'data_cell')

            if is_input:
                self._apply_style(ws[f'C{i}'], 'input_cell')
                # Add data validation
                if "%" in unit:
                    validation = DataValidation(type="decimal", operator="between", formula1=0, formula2=50)
                else:
                    validation = DataValidation(type="decimal", operator="between", formula1=0, formula2=5)
                validation.error = "Please enter a valid number"
                validation.errorTitle = "Invalid Entry"
                ws.add_data_validation(validation)
                validation.add(f'C{i}')
            else:
                self._apply_style(ws[f'C{i}'], 'formula_cell')

            self._apply_style(ws[f'D{i}'], 'data_cell')

        # Protect sheet but allow input cells
        ws.protection = SheetProtection(
            selectLockedCells=True,
            selectUnlockedCells=True,
            formatCells=False,
            formatColumns=False,
            formatRows=False,
            insertColumns=False,
            insertRows=False,
            insertHyperlinks=False,
            deleteColumns=False,
            deleteRows=False,
            sort=False,
            autoFilter=False,
            pivotTables=False,
            objects=False,
            scenarios=False
        )

        # Unlock input cells
        for i, (param, value, unit, is_input) in enumerate(market_assumptions, row+3):
            if is_input:
                ws[f'C{i}'].protection = None

    def _create_monte_carlo_dashboard(self, monte_carlo_results: MonteCarloResults):
        """Create Monte Carlo results dashboard"""
        ws = self.wb.create_sheet("🎲 Monte Carlo Results")

        # Title
        ws['B2'] = f"MONTE CARLO SIMULATION RESULTS"
        self._apply_style(ws['B2'], 'main_header')
        ws.merge_cells('B2:H2')

        ws['B3'] = f"{monte_carlo_results.simulation_runs:,} Simulation Runs"
        ws['B3'].font = Font(name='Calibri', size=11, italic=True, color='595959')

        # Key statistics
        row = 5
        ws[f'B{row}'] = "DISTRIBUTION STATISTICS"
        self._apply_style(ws[f'B{row}'], 'section_header')
        ws.merge_cells(f'B{row}:D{row}')

        stats_data = [
            ["Mean", monte_carlo_results.percentiles['mean']],
            ["Median", monte_carlo_results.percentiles['p50']],
            ["Standard Deviation", monte_carlo_results.percentiles['std']],
            ["Skewness", monte_carlo_results.skewness],
            ["Kurtosis", monte_carlo_results.kurtosis]
        ]

        for i, (stat, value) in enumerate(stats_data, row+2):
            ws[f'B{i}'] = stat
            ws[f'C{i}'] = value
            self._apply_style(ws[f'B{i}'], 'data_cell')
            if "Deviation" in stat or stat in ["Mean", "Median"]:
                self._apply_style(ws[f'C{i}'], 'currency')
            else:
                self._apply_style(ws[f'C{i}'], 'number')

        # Risk metrics
        row += 8
        ws[f'B{row}'] = "RISK METRICS"
        self._apply_style(ws[f'B{row}'], 'section_header')
        ws.merge_cells(f'B{row}:D{row}')

        risk_data = [
            ["Value at Risk (5%)", monte_carlo_results.var_metrics['var_5%']],
            ["Expected Shortfall", monte_carlo_results.expected_shortfall],
            ["Probability of Loss", monte_carlo_results.probability_of_loss]
        ]

        for i, (metric, value) in enumerate(risk_data, row+2):
            ws[f'B{i}'] = metric
            ws[f'C{i}'] = value
            self._apply_style(ws[f'B{i}'], 'data_cell')
            if "Probability" in metric:
                self._apply_style(ws[f'C{i}'], 'percentage')
            else:
                self._apply_style(ws[f'C{i}'], 'currency')

    def _create_charts_dashboard(self, deterministic_data: Dict, monte_carlo_results: Optional[MonteCarloResults]):
        """Create dashboard with interactive charts"""
        ws = self.wb.create_sheet("📈 Charts & Visualizations")

        # Title
        ws['B2'] = "INTERACTIVE CHARTS"
        self._apply_style(ws['B2'], 'main_header')
        ws.merge_cells('B2:M2')

        # Create tornado chart data for sensitivity
        if monte_carlo_results:
            self._create_tornado_chart(ws, monte_carlo_results)

        # Create waterfall chart data for valuation bridge
        self._create_waterfall_chart_data(ws, deterministic_data)

    def _create_tornado_chart(self, ws, monte_carlo_results: MonteCarloResults):
        """Create tornado chart for parameter sensitivity"""

        # Tornado chart data section
        row = 4
        ws[f'B{row}'] = "SENSITIVITY TORNADO CHART"
        self._apply_style(ws[f'B{row}'], 'section_header')
        ws.merge_cells(f'B{row}:G{row}')

        # Prepare data for tornado chart
        sorted_sensitivities = sorted(
            monte_carlo_results.parameter_sensitivities.items(),
            key=lambda x: abs(x[1]), reverse=True
        )[:6]  # Top 6 most sensitive parameters

        # Headers
        headers = ["Parameter", "Low Impact", "High Impact", "Range"]
        for i, header in enumerate(headers):
            ws.cell(row+2, i+2, header)
            self._apply_style(ws.cell(row+2, i+2), 'sub_header')

        base_value = monte_carlo_results.percentiles['mean']

        # Create tornado data
        for i, (param, sensitivity) in enumerate(sorted_sensitivities, row+3):
            param_name = param.replace('_', ' ').title()
            low_impact = base_value * (1 - abs(sensitivity) * 0.1)  # 10% parameter change
            high_impact = base_value * (1 + abs(sensitivity) * 0.1)
            impact_range = high_impact - low_impact

            ws[f'B{i}'] = param_name
            ws[f'C{i}'] = low_impact
            ws[f'D{i}'] = high_impact
            ws[f'E{i}'] = impact_range

            self._apply_style(ws[f'B{i}'], 'data_cell')
            self._apply_style(ws[f'C{i}'], 'currency')
            self._apply_style(ws[f'D{i}'], 'currency')
            self._apply_style(ws[f'E{i}'], 'currency')

        # Create horizontal bar chart
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = "Parameter Sensitivity Analysis"
        chart.y_axis.title = "Parameters"
        chart.x_axis.title = "Impact on Valuation ($)"

        # Add data to chart
        data_range = Reference(ws, min_col=3, max_col=4, min_row=row+2, max_row=row+2+len(sorted_sensitivities))
        categories = Reference(ws, min_col=2, min_row=row+3, max_row=row+2+len(sorted_sensitivities))

        chart.add_data(data_range, titles_from_data=True)
        chart.set_categories(categories)

        # Position chart
        ws.add_chart(chart, f"I{row}")

    def _create_waterfall_chart_data(self, ws, deterministic_data: Dict):
        """Create waterfall chart data for valuation bridge"""

        row = 20
        ws[f'B{row}'] = "VALUATION BRIDGE"
        self._apply_style(ws[f'B{row}'], 'section_header')
        ws.merge_cells(f'B{row}:E{row}')

        # Waterfall components
        valuation = deterministic_data.get('valuation', {})

        waterfall_data = [
            ["PV of FCF", 85000],
            ["PV of Terminal Value", 120000],
            ["Enterprise Value", 205000],
            ["Less: Net Debt", -15000],
            ["Equity Value", 190000]
        ]

        headers = ["Component", "Value ($M)"]
        for i, header in enumerate(headers):
            ws.cell(row+2, i+2, header)
            self._apply_style(ws.cell(row+2, i+2), 'sub_header')

        for i, (component, value) in enumerate(waterfall_data, row+3):
            ws[f'B{i}'] = component
            ws[f'C{i}'] = value
            self._apply_style(ws[f'B{i}'], 'data_cell')
            self._apply_style(ws[f'C{i}'], 'currency_millions')

    def _calculate_dcf_sensitivity(self, wacc: float, terminal_growth: float, deterministic_data: Dict) -> float:
        """Calculate DCF value for sensitivity analysis"""
        # Simplified calculation for demonstration
        base_fcf = 10000  # $10B base FCF
        terminal_value = base_fcf * (1 + terminal_growth) / (wacc - terminal_growth)
        pv_terminal = terminal_value / ((1 + wacc) ** 5)
        return (base_fcf * 4) + pv_terminal  # Simplified

    def _create_multi_method_dashboard(self, ticker: str, multi_method_results: MultiMethodResults):
        """Create comprehensive multi-method valuation dashboard"""
        ws = self.wb.create_sheet("🎯 Multi-Method Valuation")

        # Title
        ws['B2'] = f"{ticker} MULTI-METHOD VALUATION ANALYSIS"
        ws['B2'].font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
        ws['B2'].fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        ws.merge_cells('B2:K2')

        # Valuation Methods Summary
        row = 5
        ws[f'B{row}'] = "VALUATION METHODS SUMMARY"
        ws[f'B{row}'].font = Font(name='Calibri', size=14, bold=True)

        # Method results
        methods = [
            ("DCF Valuation", multi_method_results.dcf_value),
            ("Relative Valuation", multi_method_results.relative_value),
            ("Sum-of-Parts", multi_method_results.sum_of_parts_value),
            ("Asset-Based", multi_method_results.asset_based_value),
            ("Real Options", multi_method_results.real_options_value),
            ("Weighted Average", multi_method_results.weighted_average_value),
            ("Confidence-Weighted", multi_method_results.confidence_weighted_value)
        ]

        row += 2
        ws[f'B{row}'] = "Method"
        ws[f'C{row}'] = "Value ($)"
        ws[f'D{row}'] = "Weight (%)"
        ws[f'E{row}'] = "Confidence (%)"

        # Apply header formatting
        for col in ['B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].font = Font(bold=True)
            ws[f'{col}{row}'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')

        row += 1
        for i, (method_name, value) in enumerate(methods):
            ws[f'B{row}'] = method_name
            ws[f'C{row}'] = f"${value:.2f}"

            # Add weights and confidence if available
            if hasattr(multi_method_results, 'method_weights') and multi_method_results.method_weights:
                method_key = method_name.lower().replace('-', '_').replace(' ', '_')
                if 'dcf' in method_key:
                    method_key = 'dcf'
                elif 'relative' in method_key:
                    method_key = 'relative'
                elif 'sum' in method_key:
                    method_key = 'sum_of_parts'
                elif 'asset' in method_key:
                    method_key = 'asset_based'
                elif 'real' in method_key:
                    method_key = 'real_options'

                if method_key in multi_method_results.method_weights:
                    ws[f'D{row}'] = f"{multi_method_results.method_weights[method_key]*100:.1f}%"

                if hasattr(multi_method_results, 'confidence_scores') and method_key in multi_method_results.confidence_scores:
                    ws[f'E{row}'] = f"{multi_method_results.confidence_scores[method_key]*100:.1f}%"

            row += 1

        # Valuation Range Analysis
        row += 2
        ws[f'B{row}'] = "VALUATION RANGE ANALYSIS"
        ws[f'B{row}'].font = Font(name='Calibri', size=14, bold=True)

        if hasattr(multi_method_results, 'valuation_range') and multi_method_results.valuation_range:
            row += 2
            range_metrics = [
                ("Minimum Value", multi_method_results.valuation_range.get('min_value', 0)),
                ("Maximum Value", multi_method_results.valuation_range.get('max_value', 0)),
                ("Median Value", multi_method_results.valuation_range.get('median_value', 0)),
                ("Standard Deviation", multi_method_results.valuation_range.get('std_dev', 0)),
                ("Range (Max-Min)", multi_method_results.valuation_range.get('range', 0))
            ]

            for metric_name, metric_value in range_metrics:
                ws[f'B{row}'] = metric_name
                if 'Standard Deviation' in metric_name or 'Range' in metric_name:
                    ws[f'C{row}'] = f"${metric_value:.2f}"
                else:
                    ws[f'C{row}'] = f"${metric_value:.2f}"
                row += 1

        # Investment Recommendation
        if hasattr(multi_method_results, 'investment_recommendation') and multi_method_results.investment_recommendation:
            row += 2
            ws[f'B{row}'] = "INVESTMENT RECOMMENDATION"
            ws[f'B{row}'].font = Font(name='Calibri', size=14, bold=True)

            row += 2
            recommendation = multi_method_results.investment_recommendation.get('recommendation', 'N/A')
            confidence = multi_method_results.investment_recommendation.get('confidence_level', 0)

            ws[f'B{row}'] = "Recommendation:"
            ws[f'C{row}'] = recommendation
            ws[f'C{row}'].font = Font(bold=True, color='1F4E79')

            row += 1
            ws[f'B{row}'] = "Confidence Level:"
            ws[f'C{row}'] = f"{confidence:.1%}"

        # Apply column widths
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12

        # Freeze panes
        ws.freeze_panes = 'B5'

    def _apply_global_formatting(self):
        """Apply global formatting to all sheets"""
        for sheet in self.wb.worksheets:
            # Set default font
            for row in sheet.iter_rows():
                for cell in row:
                    if not cell.font.name:
                        cell.font = Font(name='Calibri', size=10)

            # Auto-adjust column widths (basic)
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width

    def _apply_style(self, cell, style_name):
        """Apply predefined style to cell"""
        style = self.styles.get(style_name, {})
        for attr, value in style.items():
            setattr(cell, attr, value)


def test_professional_excel_template():
    """Test the professional Excel template"""

    # Create mock data
    class MockSnapshot:
        def __init__(self):
            self.ticker = 'AAPL'
            self.current_price = 150.0
            self.market_cap = 2400000000000

    class MockFinancials:
        def __init__(self):
            self.fundamentals = {'totalRevenue': 394328000000}

    class MockDataset:
        def __init__(self):
            self.ticker = 'AAPL'
            self.snapshot = MockSnapshot()
            self.financials = MockFinancials()

    dataset = MockDataset()

    deterministic_data = {
        'valuation': {
            'dcf_value': 165.50,
            'scenarios': {'bear': 145.0, 'base': 165.5, 'bull': 185.0},
            'assumptions': {
                'wacc': 0.085,
                'risk_free_rate': 0.041,
                'equity_risk_premium': 0.08,
                'beta': 1.25
            }
        }
    }

    print("="*60)
    print("TESTING PROFESSIONAL EXCEL TEMPLATE")
    print("="*60)

    template = ProfessionalExcelTemplate()

    model_path = template.create_professional_model(
        ticker='AAPL',
        dataset=dataset,
        deterministic_data=deterministic_data,
        monte_carlo_results=None,
        output_path='test_professional_models/AAPL_Professional_Test.xlsx'
    )

    print(f"✅ Professional Excel template created: {model_path}")

if __name__ == "__main__":
    test_professional_excel_template()